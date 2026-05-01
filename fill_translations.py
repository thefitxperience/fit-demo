"""
Fill TranslationTemplateFoodDish.xlsx and TranslationTemplateIngredient.xlsx
with Arabic translations sourced from translations-ar.json and a comprehensive
ingredient dictionary.
"""
import json, re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

# ── 1. Load existing JSON translations ───────────────────────────────────────
with open("assets/translations-ar.json", encoding="utf-8") as f:
    ar = json.load(f)

FOOD_AR   = ar.get("food", {})        # English name  → Arabic name
DESC_AR   = ar.get("descriptions", {})  # English desc  → Arabic desc
ING_AR    = ar.get("ingredients", {})   # English name  → Arabic name

# ── 2. Comprehensive ingredient translation dictionary ────────────────────────
# Covers all 314 ingredients in the entity export.
EXTRA_ING = {
    # Breads & Grains (IDs 1-29, 240-254…)
    "Pita Whole Wheat Bread":                  "خبز بيتا من القمح الكامل",
    "Baguette, multigrain":                    "باغيت متعدد الحبوب",
    "Bagel Multigrain":                        "بيغل متعدد الحبوب",
    "Kaak, Round":                             "كعك دائري",
    "Markouk bread":                           "خبز مرقوق",
    "Homemade Pancake":                        "فطيرة منزلية",
    "Rice Cake, Whole Grain":                  "كعكة الأرز الكاملة",
    "Rice Cake, Multigrain":                   "كعكة الأرز متعددة الحبوب",
    "Rice Cake, Dark Chocolate Coated":        "كعكة الأرز مغطاة بالشوكولاته الداكنة",
    "Toast Dry":                               "توست جاف",
    "Toast Whole grains, Soft":                "توست حبوب كاملة طري",
    "Freekeh, boiled":                         "فريكة مسلوقة",
    "Bulgur, boiled":                          "برغل مسلوق",
    "Granola":                                 "غرانولا",
    "Muesli":                                  "موسلي",
    "Lupine":                                  "ترمس",
    "Oats, Dry Gluten Free":                   "شوفان جاف خالي من الغلوتين",
    "Oats, Dry":                               "شوفان جاف",
    "Pasta, Wheat, boiled":                    "معكرونة قمح مسلوقة",
    "Pasta, Wheat, Whole grain, Boiled":       "معكرونة قمح كاملة مسلوقة",
    "Pasta, Gluten Free, boiled":              "معكرونة خالية من الغلوتين مسلوقة",
    "Rice, boiled":                            "أرز مسلوق",
    "Rice Brown, Whole grain, boiled":         "أرز بني كامل الحبة مسلوق",
    "Couscous, boiled":                        "كسكس مسلوق",
    "Quinoa (cooked, plain)":                  "كينوا مطبوخة سادة",
    "Corn, canned":                            "ذرة معلبة",
    "Chestnuts":                               "كستناء",
    "Green peas, cooked":                      "بازلاء مطبوخة",
    "Potatoes baked or boiled":                "بطاطس مشوية أو مسلوقة",
    "Sweet potatoes, baked or boiled":         "بطاطا حلوة مشوية أو مسلوقة",
    "Kidney Beans, cooked or canned":          "فاصولياء حمراء مطبوخة أو معلبة",
    "Lima Beans, cooked or canned":            "فاصولياء ليما مطبوخة أو معلبة",
    "Chickpeas Beans, cooked or canned":       "حمص مطبوخ أو معلب",
    "Pinto Beans, cooked or canned":           "فاصولياء بينتو مطبوخة أو معلبة",
    "Lentils, cooked":                         "عدس مطبوخ",
    "Edamame":                                 "إيدامامي",
    "Mixed Nuts (Salted)":                     "مكسرات مشكلة (مالحة)",
    "Whole-Grain Crackers":                    "بسكويت الحبوب الكاملة",
    "Roasted Chickpeas":                       "حمص محمص",
    "Seaweed Snacks":                          "وجبة خفيفة من الطحالب",
    "Corn Tortilla":                           "تورتيا الذرة",
    "Low-Carb Tortilla":                       "تورتيا قليلة الكربوهيدرات",
    "Taco Shell (Hard, Corn)":                 "قشرة تاكو (صلبة، ذرة)",
    "Taco Shell (Baked, Corn)":                "قشرة تاكو (مخبوزة، ذرة)",
    "Chapati":                                 "خبز الشاباتي",
    "Oats, Cooked or Dry Gluten Free":         "شوفان مطبوخ أو جاف خالي من الغلوتين",
    "Oats, Cooked or Dry":                     "شوفان مطبوخ أو جاف",
    "Buckwheat (cooked)":                      "حنطة السوداء مطبوخة",
    "Yellow moong Dal cooked":                 "عدس أصفر موونج دال مطبوخ",

    # Fruits
    "Apple, Dried":                            "تفاح مجفف",
    "Apple, unpeeled, small":                  "تفاحة صغيرة بقشرها",
    "Apricot, fresh":                          "مشمش طازج",
    "Apricots Dried":                          "مشمش مجفف",
    "Banana":                                  "موز",
    "Blackberries, Fresh or Frozen":           "توت أسود طازج أو مجمد",
    "Blueberries, Dried":                      "توت أزرق مجفف",
    "Blueberries, Fresh or Frozen":            "توت أزرق طازج أو مجمد",
    "Carrot Juice":                            "عصير الجزر",
    "Cherries, Dried":                         "كرز مجفف",
    "Cherries, fresh":                         "كرز طازج",
    "Cranberries, Dried":                      "توت القرابيا المجفف",
    "Custard Apple \u201cKashta\u201d":        "القشطة",
    "Dates (Deglet Noor)":                     "تمر دقلة النور",
    "Dates (Medjool)":                         "تمر المجدول",
    "Dried Sweetened Mango":                   "مانغو مجففة محلاة",
    "Figs Dried":                              "تين مجفف",
    "Figs, fresh":                             "تين طازج",
    "Goji Berries":                            "توت غوجي",
    "Grapefruit":                              "جريب فروت",
    "Grapefruit Juice, fresh":                 "عصير جريب فروت طازج",
    "Grapes":                                  "عنب",
    "Guava":                                   "جوافة",
    'Jujuba "Ennab"':                          "عناب",
    "Kiwi":                                    "كيوي",
    "Loquat":                                  "إكيدنيا",
    "Mango":                                   "مانغو",
    "Nectarine":                               "نكتارين",
    "Orange":                                  "برتقال",
    "Papaya":                                  "بابايا",
    "Pear Dried":                              "كمثرى مجففة",
    "Pear Fresh":                              "كمثرى طازجة",
    'Persimmon "Kaki\u201d':                   "كاكا (فاكهة الكاكي)",
    "Pineapple, Dried":                        "أناناس مجفف",
    "Pineapple, Fresh":                        "أناناس طازج",
    "Plum":                                    "برقوق",
    "Pomegranate":                             "رمان",
    "Pomegranate Juice":                       "عصير الرمان",
    "Pomelo":                                  "بوملي",
    "Prickly Pear":                            "تين شوكي",
    "Prunes Dried":                            "برقوق مجفف",
    "Raisins, Drie":                           "زبيب",
    "Raspberries, Fresh or Frozen":            "توت العليق طازج أو مجمد",
    "Sour Plums":                              "جانريك (برقوق حامض)",
    "Strawberries":                            "فراولة",
    "Tamarinds":                               "تمرهندي",
    "Tomato Juice":                            "عصير الطماطم",
    "Watermelon":                              "بطيخ",
    "Yellow Melon":                            "شمام أصفر",

    # Vegetables
    "Artichoke cooked":                        "أرضي شوكي مطبوخ",
    "Asparagus Greens, cooked":                "هليون مطبوخ",
    "Beets, cooked":                           "بنجر مطبوخ",
    "Broccoli Raw":                            "بروكلي طازج",
    "Broccoli Cooked":                         "بروكلي مطبوخ",
    "Brussels sprouts":                        "كرنب بروكسل",
    "Cabbage":                                 "ملفوف",
    "Cauliflower":                             "قرنبيط",
    "Eggplant raw":                            "باذنجان طازج",
    "Eggplant cooked":                         "باذنجان مطبوخ",
    "Leeks cooked":                            "كراث مطبوخ",
    "Mushrooms Cooked":                        "فطر مطبوخ",
    "Okra":                                    "بامية",
    "Okra cooked":                             "بامية مطبوخة",
    "Onions Raw":                              "بصل طازج",
    "Onions Cooked":                           "بصل مطبوخ",
    "Peppers (green), cooked":                 "فلفل أخضر مطبوخ",
    "Romaine":                                 "خس روماني",
    "Rutabaga":                                "لفت أصفر",
    "Salad Greens and lettuce":                "خضار الخس والسلطة",
    "Sauerkraut":                              "مخلل الملفوف الألماني",
    "Spinach, Raw":                            "سبانخ طازجة",
    "Spinach, cooked":                         "سبانخ مطبوخة",
    "Summer squash cooked":                    "كوسا صيفية مطبوخة",
    "Tomato":                                  "طماطم",
    "Turnip, cooked":                          "لفت مطبوخ",
    "Zucchini cooked":                         "كوسا مطبوخة",
    "Cucumber":                                "خيار",
    "Kale, raw":                               "كيل طازج",
    "Carrot, cooked":                          "جزر مطبوخ",
    "Kimchi":                                  "كيمتشي",

    # Dairy & Alternatives
    "Milk Powder - Low fat":                   "حليب بودرة منخفض الدسم",
    "Milk, half Skimmed":                      "حليب نصف خالي الدسم",
    "Milk, Fat Free":                          "حليب خالي الدسم",
    "Milk, Lactose Free Half Skimmed":         "حليب خالي اللاكتوز نصف خالي الدسم",
    "Milk, Full Fat":                          "حليب كامل الدسم",
    "Yogurt Low Fat":                          "لبن قليل الدسم",
    "Yogurt, Lactose Free":                    "لبن خالي اللاكتوز",
    "Yogurt Full Fat":                         "لبن كامل الدسم",
    "Greek Yogurt 0% Fat":                     "لبن يوناني خالي الدسم",
    "Greek Yogurt Full Fat":                   "لبن يوناني كامل الدسم",
    "Cheese (Double cr\u00e8me cheese/low fat white cheese)": "جبنة (جبنة كريمية مزدوجة / جبنة بيضاء قليلة الدسم)",
    "Cheese Akkawi":                           "جبنة عكاوي",
    "Cheese Brie":                             "جبنة بري",
    "Cheese Feta":                             "جبنة فيتا",
    "Cheese Halloum":                          "جبنة حلوم",
    "Cheese Mozarella Low fat":                "جبنة موزاريلا قليلة الدسم",
    "Cheese Mozzarella":                       "جبنة موزاريلا",
    "Cheese Swiss":                            "جبنة سويسرية",
    "Cheese Parmesan":                         "جبنة بارميزان",
    "Cheddar":                                 "جبنة شيدر",
    "Labneh":                                  "لبنة",
    "Labneh, low fat":                         "لبنة قليلة الدسم",
    "Cottage Cheese (Full-Fat)":               "جبنة القريش كاملة الدسم",
    "Cottage Cheese (Low-Fat)":                "جبنة القريش قليلة الدسم",
    "Ricotta (Whole Milk)":                    "ريكوتا (حليب كامل)",
    "Ricotta (Part-Skim)":                     "ريكوتا (نصف خالي الدسم)",
    "Cream Cheese (Light)":                    "جبنة كريمية خفيفة",
    "Blue Cheese (Gorgonzola)":                "جبنة زرقاء (غورغونزولا)",
    "Paneer":                                  "جبنة بانير",
    "Queso Fresco":                            "جبنة كيسو فريسكو",
    "Gouda":                                   "جبنة غودة",
    "Buttermilk (Traditional)":               "لبن الزبدة التقليدي",
    "Goat\u2019s Milk":                        "حليب الماعز",
    "Sheep\u2019s Milk":                       "حليب الأغنام",
    "Skyr":                                    "سكير",
    "Kefir (Plain)":                           "كفير سادة",
    "Frozen Yogurt (Plain)":                   "لبن مثلج سادة",
    "Kashkaval Cheese":                        "جبنة كشكافال",
    "Jameed (Dried Yogurt)":                   "جميد (لبن جاف)",
    "Almond Milk, unsweetened":                "حليب اللوز غير المحلى",
    "Oat Milk, unsweetened":                   "حليب الشوفان غير المحلى",
    "Coconut milk, Unsweetened":               "حليب جوز الهند غير المحلى",
    "Cashew Milk, Unsweetened":                "حليب الكاشو غير المحلى",
    "Vegan Cheese":                            "جبنة نباتية",

    # Proteins – Animal
    "Chicken Breast, without skin, cooked/grilled": "صدر الدجاج بدون جلد مطبوخ/مشوي",
    "Turkey Breast, without skin, cooked/grilled":  "صدر الديك الرومي بدون جلد مطبوخ/مشوي",
    "Chicken thigh, without skin, cooked/grilled":  "فخذ الدجاج بدون جلد مطبوخ/مشوي",
    "Turkey Bacon":                            "لحم مقدد من الديك الرومي",
    "Egg, white, boiled":                      "بياض البيض مسلوق",
    "Egg, white, scrambled (no fat added)":    "بياض البيض مخفوق (بدون دهن)",
    "Egg, white, fried (with 1 tsp oil)":      "بياض البيض مقلي (بملعقة زيت)",
    "Egg, whole, boiled":                      "بيض مسلوق",
    "Egg, whole, Scrambled (no fat added)":    "بيض مخفوق بدون دهن",
    "Egg, whole, Fried (with 1 tsp oil)":      "بيض مقلي (بملعقة زيت)",
    "Fish, Salmon, Raw (sashimi)":             "سمك السلمون طازج (ساشيمي)",
    "Fish, Salmon, Grilled":                   "سمك السلمون مشوي",
    "Fish, Tuna, Raw (sashimi)":               "تونة طازجة (ساشيمي)",
    "Fish, Tuna, Grilled":                     "تونة مشوية",
    "Fish, Cod, Raw (sashimi)":                "سمك الكود طازج (ساشيمي)",
    "Fish, Cod, Grilled":                      "سمك الكود مشوي",
    "Hamburger 90% Lean Meat, Patty, Grilled": "باتي برغر لحم قليل الدهن 90%، مشوي",
    "Grounded Chicken, Patty, Grilled":        "باتي دجاج مفروم مشوي",
    "Grounded Chicken, Patty, Breaded":        "باتي دجاج مفروم مع البانيه",
    "Lean meat stew, grilled":                 "يخنة لحم قليل الدهن مشوية",
    "Chicken breast skeweres, Taouk, grilled": "أسياخ صدر دجاج (تاووك) مشوي",
    "Lean minced meat, cooked without oil":    "لحم مفروم قليل الدهن مطبوخ بدون زيت",
    "Sardine in oil, drained":                 "سردين بالزيت مصفى",
    "Shellfish Crab, grilled":                 "سلطعون مشوي",
    "Shellfish Shrimps, grilled":              "روبيان مشوي",
    "Fish, Tuna, Canned in water, drained":    "تونة معلبة بالماء مصفاة",
    "Fish, Tuna, Canned in Oil, drained":      "تونة معلبة بالزيت مصفاة",

    # Proteins – Plant
    "Tempeh, Raw":                             "تمبيه خام",
    "Tempeh, Grilled":                         "تمبيه مشوي",
    "Tofu, Cooked":                            "توفو مطبوخ",
    "Tofu, Raw Firm":                          "توفو خام صلب",
    "Tofu, Raw":                               "توفو خام",
    "Tofu, Grilled":                           "توفو مشوي",
    "Falafel, Baked with oil":                 "فلافل بالفرن مع زيت",
    "Vegan Burger Patty":                      "باتي برغر نباتي",
    "Vegan Protein Powder":                    "مسحوق بروتين نباتي",

    # Fats & Spreads
    "Almond butter":                           "زبدة اللوز",
    "Peanut Butter":                           "زبدة الفول السوداني",
    "Avocado":                                 "أفوكادو",
    "Butter, Reduced-fat":                     "زبدة قليلة الدهون",
    "Cream cheese, Reduced-fat":               "جبنة كريمية قليلة الدهون",
    "Mayonnaise, Reduced-fat":                 "مايونيز منخفض الدهون",
    "Almonds":                                 "لوز",
    "Cashews":                                 "كاشو",
    "Hazelnuts":                               "بندق",
    "Peanuts":                                 "فول سوداني",
    "Walnuts":                                 "جوز",
    "Pecans":                                  "جوز البيكان",
    "Pistachios":                              "فستق",
    "Pine nuts":                               "صنوبر",
    "Oil, olive":                              "زيت الزيتون",
    "Olives, black or green":                  "زيتون أسود أو أخضر",
    "Pumpkin Seeds":                           "بذور اليقطين",
    "Sesame Seeds":                            "بذور السمسم",
    "Sunflower Seeds":                         "بذور عباد الشمس",
    "Flaxeed Seeds":                           "بذور الكتان",
    "Tahini/sesame paste":                     "طحينة",
    "Clarified Butter (Ghee)":                 "سمن مصفى",
    "Margarine (Plant-Based)":                 "مارغرين نباتي",
    "Dark Chocolate (70%)":                    "شوكولاته داكنة (70%)",
    "Fruit & Nut Trail Mix":                   "خليط الفواكه والمكسرات",
    "Brazil Nuts":                             "جوز البرازيل",
    "Chia Seeds":                              "بذور الشيا",

    # Mojibake variants (xlsx stored with wrong encoding — exact strings from file)
    "Cheese (Double cr\u221a\u00aeme cheese/low fat white cheese)": "جبنة (جبنة كريمية مزدوجة / جبنة بيضاء قليلة الدسم)",
    "Goat\u201a\u00c4\u00f4s Milk":            "حليب الماعز",
    "Sheep\u201a\u00c4\u00f4s Milk":           "حليب الأغنام",
    "Custard Apple \u201a\u00c4\u00faKashta\u201a\u00c4\u00f9": "القشطة",
    'Persimmon "Kaki\u201a\u00c4\u00f9':       "كاكا (فاكهة الكاكي)",

    # Beverages
    "Tea, Instant, Decaf, Unsweetened":        "شاي فوري منزوع الكافيين غير محلى",
    "Black Coffee (Brewed)":                   "قهوة سوداء (مطبوخة)",
    "Espresso (Single Shot)":                  "إسبريسو (جرعة مفردة)",
    "Coffee with Skim Milk":                   "قهوة مع حليب خالي الدسم",
    "Carbonated Water (Plain)":                "ماء غازي سادة",
    "Sparkling Flavored Water (Unsweetened)":  "ماء فوار منكّه غير محلى",
    "Sparkling Flavored Water (Sweetened)":    "ماء فوار منكّه محلى",
    "Hot Chocolate (Powder + Water)":          "شوكولاته ساخنة (مسحوق + ماء)",
    "Matcha Latte (Unsweetened)":              "ماتشا لاتيه (غير محلى)",
    "Matcha Latte, Lactose Free (Unsweetened)":"ماتشا لاتيه خالي اللاكتوز (غير محلى)",
    "Chai Tea Latte (Unsweetened)":            "شاي تشاي لاتيه (غير محلى)",
    "Chai Tea Latte, Lactose Free (Unsweetened)": "شاي تشاي لاتيه خالي اللاكتوز (غير محلى)",
    "Almond Milk Latte (Unsweetened)":         "لاتيه حليب اللوز (غير محلى)",
    "Oat Milk Latte (Unsweetened)":            "لاتيه حليب الشوفان (غير محلى)",

    # Condiments & Misc
    "Hummus":                                  "حمص",
    "Baba Ghanoush":                           "بابا غنوج",
    "Lentil Soup":                             "شوربة العدس",
}

# Merge: EXTRA_ING takes base, then ING_AR (JSON) overrides anything it has
ALL_ING = {**EXTRA_ING, **ING_AR}

# ── 3. Helper: clean encoding artefacts in entity-export text ─────────────────
def clean(txt):
    if not txt:
        return ""
    # Fix common mojibake sequences that appear in the xlsx
    txt = txt.replace("saut\u221a\u00a9ed", "saut\u00e9ed")
    txt = txt.replace("\u221a\u00a9", "\u00e9")
    txt = txt.replace("cr\u221a\u00e8me", "cr\u00e8me")
    txt = txt.replace("Goat\u2030\u00c4\u00f4s", "Goat\u2019s")
    txt = txt.replace("Sheep\u2030\u00c4\u00f4s", "Sheep\u2019s")
    txt = txt.replace("\u2030\u00c4\u00f4", "\u2019")
    return txt.strip()

# ── 4. Load entity exports ────────────────────────────────────────────────────
wb_dish = openpyxl.load_workbook("EntityExport_20260414_1851.xlsx")
ws_dish = wb_dish.active
dishes = []   # list of (id, en_name, en_desc)
for row in ws_dish.iter_rows(min_row=2, values_only=True):
    fid, fname, fdesc, _ = row
    dishes.append((fid, clean(str(fname)), clean(str(fdesc) if fdesc else "")))

wb_ing = openpyxl.load_workbook("EntityExport_20260414_1852 (1).xlsx")
ws_ing = wb_ing.active
ingredients = []   # list of (id, en_name)
for row in ws_ing.iter_rows(min_row=2, values_only=True):
    iid, iname = row
    if isinstance(iname, str):
        ingredients.append((iid, clean(iname)))
    # skip rows where ingredient name is numeric (data error in entity export)

# ── 5. Build description lookup: English name → Arabic description ────────────
# Match by English name (unique per dish) via the entity-export description field.
# We first build: en_desc → ar_desc from JSON, then map via en_name → en_desc.
name_to_ar_desc = {}
for (fid, fname, fdesc) in dishes:
    ar_desc = DESC_AR.get(fdesc, "")
    name_to_ar_desc[fname] = ar_desc

# ── 6. Fill TranslationTemplateFoodDish.xlsx ──────────────────────────────────
wb_t = openpyxl.load_workbook("TranslationTemplateFoodDish.xlsx")
ws_t = wb_t.active

# Style the header row
header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws_t[1]:
    if cell.value:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Write data rows
for i, (fid, fname, fdesc) in enumerate(dishes, start=2):
    ar_name = FOOD_AR.get(fname, "")
    ar_desc = DESC_AR.get(fdesc, "")
    ar_prep = ""  # all preparationInfo values are null in entity export

    ws_t.cell(row=i, column=1, value=fid)
    ws_t.cell(row=i, column=2, value=ar_name)
    ws_t.cell(row=i, column=3, value=ar_desc)
    ws_t.cell(row=i, column=4, value=ar_prep)

    # Style
    ws_t.cell(row=i, column=1).alignment = Alignment(horizontal="center")
    for col in (2, 3, 4):
        cell = ws_t.cell(row=i, column=col)
        cell.alignment = Alignment(horizontal="right", vertical="center",
                                   wrap_text=True, readingOrder=2)

# Column widths
ws_t.column_dimensions["A"].width = 12
ws_t.column_dimensions["B"].width = 40
ws_t.column_dimensions["C"].width = 70
ws_t.column_dimensions["D"].width = 40

wb_t.save("TranslationTemplateFoodDish.xlsx")
print(f"✓ TranslationTemplateFoodDish.xlsx filled — {len(dishes)} dishes")

# Check coverage
missing_names = [fname for (_, fname, _) in dishes if not FOOD_AR.get(fname)]
missing_descs = [(fname, fdesc) for (_, fname, fdesc) in dishes if fdesc and not DESC_AR.get(fdesc)]
if missing_names:
    print(f"  ⚠ Missing Arabic name for: {missing_names}")
if missing_descs:
    print(f"  ⚠ Missing Arabic description for:")
    for n, d in missing_descs:
        print(f"     [{n}] → \"{d}\"")

# ── 7. Fill TranslationTemplateIngredient.xlsx ────────────────────────────────
wb_i = openpyxl.load_workbook("TranslationTemplateIngredient.xlsx")
ws_i = wb_i.active

# Style header
for cell in ws_i[1]:
    if cell.value:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

missing_ing = []
for i, (iid, iname) in enumerate(ingredients, start=2):
    ar_name = ALL_ING.get(iname, "")
    if not ar_name:
        missing_ing.append(iname)

    ws_i.cell(row=i, column=1, value=iid)
    ws_i.cell(row=i, column=2, value=ar_name)

    ws_i.cell(row=i, column=1).alignment = Alignment(horizontal="center")
    ws_i.cell(row=i, column=2).alignment = Alignment(
        horizontal="right", vertical="center", wrap_text=True, readingOrder=2)

ws_i.column_dimensions["A"].width = 14
ws_i.column_dimensions["B"].width = 55

wb_i.save("TranslationTemplateIngredient.xlsx")
print(f"✓ TranslationTemplateIngredient.xlsx filled — {len(ingredients)} ingredients")
if missing_ing:
    print(f"  ⚠ No Arabic translation found for {len(missing_ing)} ingredient(s):")
    for n in missing_ing:
        print(f"     \"{n}\"")
else:
    print("  ✓ All ingredients translated")
